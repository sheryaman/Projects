import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook , load_workbook
import re
import os

nombre_archivo = "datos.xlsx"
if os.path.exist(nombre_archivo):
   wb = load_workbook(nombre_archivo)
   ws = wb.active
else:

 wb = Workbook()
 ws = wb.active
 ws.append(["Nombre","Edad","Email","Telefono","Dirección"])
 wb.save("datos.xlsx")
root = tk.Tk()


def guardar_datos():
    nombre = entry_nombre.get()
    edad = Entry_edad.get()
    email = Entry_email.get()
    telefono = entry_telefono.get()
    direccion = entry_direccion.get()
    
    if not nombre or not edad or not email or not  telefono or not direccion:
        messagebox.showwarning(title="Advertencia",message="Campo Obligatorio")
        return
    try:
        edad = int(edad)
        telefono = int(telefono)
    except ValueError:
        messagebox.showwarning(title="Error",message="Introducir Variables Numericas en los campos requeridos")
        return
    
    if not re.match(r"[^@]+@[^@]+\.[^@]+" , email):
        messagebox.showwarning(title="Advertencia",message="Introducir un  correo valido")
        return
    ws.append([nombre,edad,email,telefono,direccion])
    wb.save("datos.xlsx")
    messagebox.showinfo(title="Información",message="Datos Guardados Correctamente")
    entry_nombre.delete(0,tk.END)
    Entry_edad.delete(0,tk.END)
    entry_direccion.delete(0,tk.END)
    entry_telefono.delete(0,tk.END)
    Entry_email.delete(0,tk.END)


    



root.title("Formulario De Datos")
root.configure(bg="#4B6587")
label_style = {"bg": "#4B6587" , "fg": "white" }
entry_style = {"bg": "#B3A8A8" , "fg": "black"}
label_nombre = tk.Label(root, text="Nombre", **label_style)
label_nombre.grid(row=0,column=0,padx=10,pady=5)
entry_nombre = tk.Entry(root,**entry_style)
entry_nombre.grid(row=0,column=1,padx=10,pady=5)
label_edad = tk.Label(root,text="Edad",  **label_style)
label_edad.grid(row=1,column=0,padx=10,pady=5)
Entry_edad = tk.Entry(root,**entry_style)
Entry_edad.grid(row=1,column=1,padx=10,pady=5)
label_email = tk.Label(root,text="Email",  **label_style)
label_email.grid(row=2,column=0,padx=10,pady=5)
Entry_email = tk.Entry(root,**entry_style)
Entry_email.grid(row=2,column=1,padx=10,pady=5)
label_telefono = tk.Label(root, text="Telefono", **label_style)
label_telefono.grid(row=3,column=0,padx=10,pady=5)
entry_telefono = tk.Entry(root,**entry_style)
entry_telefono.grid(row=3,column=1,padx=10,pady=5)
label_direccion = tk.Label(root, text="Dirección", **label_style)
label_direccion.grid(row=4,column=0,padx=10,pady=5)
entry_direccion = tk.Entry(root,**entry_style)
entry_direccion.grid(row=4,column=1,padx=10,pady=5)

boton_guardar = tk.Button(root,text="Guardar",command=guardar_datos, bg="#6D8299",fg="white" , width=50)
boton_guardar.grid(row=5,column=0,padx=10,pady=10, columnspan=2)




root.mainloop()